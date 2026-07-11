"""
Build excel_dashboard/Financial_Analysis_Dashboard.xlsx
Raw data + formula-driven pivot summaries + native Excel charts,
mirroring the same 6 analyses as the Python/SQL pipeline.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

SRC_CSV = "data/cleaned_financial_data.csv"
OUT_XLSX = "excel_dashboard/Financial_Analysis_Dashboard.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

df = pd.read_csv(SRC_CSV, keep_default_na=False, na_values=[""])
band_order = ["None", "Low", "Medium", "High"]

wb = Workbook()

# ---------------------------------------------------------------- Raw Data
ws_raw = wb.active
ws_raw.title = "Raw Data"
cols = list(df.columns)
for j, c in enumerate(cols, start=1):
    cell = ws_raw.cell(row=1, column=j, value=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
for i, row in enumerate(df.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws_raw.cell(row=i, column=j, value=val).font = BODY_FONT
n_rows = len(df) + 1
for j, c in enumerate(cols, start=1):
    ws_raw.column_dimensions[get_column_letter(j)].width = max(12, len(c) + 2)
ws_raw.freeze_panes = "A2"

def col_letter(name):
    return get_column_letter(cols.index(name) + 1)

SEG_COL, SALES_COL, PROFIT_COL, BAND_COL = (
    col_letter("Segment"), col_letter("Sales"), col_letter("Profit"), col_letter("Discount Band")
)
PROD_COL, COUNTRY_COL, MONTHNAME_COL, MONTHNUM_COL, YEAR_COL = (
    col_letter("Product"), col_letter("Country"), col_letter("Month Name"),
    col_letter("Month Number"), col_letter("Year")
)
RAW_RANGE = f"'Raw Data'!${{col}}$2:${{col}}${n_rows}"

def rng(col):
    return f"'Raw Data'!${col}$2:${col}${n_rows}"


def style_table_header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def style_table_body(ws, r1, r2, start_col, end_col):
    for r in range(r1, r2 + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER


def autosize(ws, ncols, min_w=12):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min_w


# ---------------------------------------------------------- 1. Segment Summary
ws = wb.create_sheet("Segment Summary")
ws["A1"] = "Profitability by Segment"
ws["A1"].font = TITLE_FONT
headers = ["Segment", "Sales ($)", "Profit ($)", "Margin (%)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=3, column=j, value=h)
style_table_header(ws, 3, 1, 4)

segments = sorted(df["Segment"].unique())
for i, seg in enumerate(segments, start=4):
    ws.cell(row=i, column=1, value=seg)
    ws.cell(row=i, column=2, value=f'=SUMIF({rng(SEG_COL)},A{i},{rng(SALES_COL)})')
    ws.cell(row=i, column=3, value=f'=SUMIF({rng(SEG_COL)},A{i},{rng(PROFIT_COL)})')
    ws.cell(row=i, column=4, value=f'=C{i}/B{i}')
last = 3 + len(segments)
style_table_body(ws, 4, last, 1, 4)
for r in range(4, last + 1):
    ws.cell(row=r, column=2).number_format = '$#,##0'
    ws.cell(row=r, column=3).number_format = '$#,##0;($#,##0)'
    ws.cell(row=r, column=4).number_format = '0.0%'
autosize(ws, 4, 16)

chart = BarChart()
chart.type = "bar"
chart.title = "Total Profit by Segment"
chart.y_axis.title = "Segment"
chart.x_axis.title = "Profit ($)"
data = Reference(ws, min_col=3, min_row=3, max_row=last)
cats = Reference(ws, min_col=1, min_row=4, max_row=last)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height, chart.width = 9, 16
ws.add_chart(chart, "F3")

# ---------------------------------------------------- 2. Discount Band Summary
ws2 = wb.create_sheet("Discount Band Summary")
ws2["A1"] = "Profit Margin % by Discount Band"
ws2["A1"].font = TITLE_FONT
for j, h in enumerate(["Discount Band", "Sales ($)", "Profit ($)", "Margin (%)"], start=1):
    ws2.cell(row=3, column=j, value=h)
style_table_header(ws2, 3, 1, 4)
for i, band in enumerate(band_order, start=4):
    ws2.cell(row=i, column=1, value=band)
    ws2.cell(row=i, column=2, value=f'=SUMIF({rng(BAND_COL)},A{i},{rng(SALES_COL)})')
    ws2.cell(row=i, column=3, value=f'=SUMIF({rng(BAND_COL)},A{i},{rng(PROFIT_COL)})')
    ws2.cell(row=i, column=4, value=f'=C{i}/B{i}')
last2 = 3 + len(band_order)
style_table_body(ws2, 4, last2, 1, 4)
for r in range(4, last2 + 1):
    ws2.cell(row=r, column=2).number_format = '$#,##0'
    ws2.cell(row=r, column=3).number_format = '$#,##0;($#,##0)'
    ws2.cell(row=r, column=4).number_format = '0.0%'
autosize(ws2, 4, 16)

chart2 = BarChart()
chart2.type = "col"
chart2.title = "Profit Margin % by Discount Band"
chart2.y_axis.title = "Margin (%)"
chart2.x_axis.title = "Discount Band"
data2 = Reference(ws2, min_col=4, min_row=3, max_row=last2)
cats2 = Reference(ws2, min_col=1, min_row=4, max_row=last2)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height, chart2.width = 9, 16
ws2.add_chart(chart2, "F3")

# ------------------------------------------------ 3. Segment x Discount Heatmap
ws3 = wb.create_sheet("Segment x Discount Heatmap")
ws3["A1"] = "Profit Margin % — Segment vs Discount Band"
ws3["A1"].font = TITLE_FONT
ws3.cell(row=3, column=1, value="Segment")
for j, band in enumerate(band_order, start=2):
    ws3.cell(row=3, column=j, value=band)
style_table_header(ws3, 3, 1, 1 + len(band_order))
for i, seg in enumerate(segments, start=4):
    ws3.cell(row=i, column=1, value=seg)
    for j, band in enumerate(band_order, start=2):
        col_l = get_column_letter(j)
        f = (f'=SUMIFS({rng(PROFIT_COL)},{rng(SEG_COL)},$A{i},{rng(BAND_COL)},{col_l}$3)'
             f'/SUMIFS({rng(SALES_COL)},{rng(SEG_COL)},$A{i},{rng(BAND_COL)},{col_l}$3)')
        ws3.cell(row=i, column=j, value=f)
last3 = 3 + len(segments)
style_table_body(ws3, 4, last3, 1, 1 + len(band_order))
for r in range(4, last3 + 1):
    for c in range(2, 2 + len(band_order)):
        ws3.cell(row=r, column=c).number_format = '0.0%'
autosize(ws3, 1 + len(band_order), 14)

heat_range = f"B4:{get_column_letter(1 + len(band_order))}{last3}"
rule = ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B",
)
ws3.conditional_formatting.add(heat_range, rule)

# --------------------------------------------------------- 4. Product Summary
ws4 = wb.create_sheet("Product Summary")
ws4["A1"] = "Sales & Profit by Product"
ws4["A1"].font = TITLE_FONT
for j, h in enumerate(["Product", "Sales ($)", "Profit ($)", "Margin (%)"], start=1):
    ws4.cell(row=3, column=j, value=h)
style_table_header(ws4, 3, 1, 4)
products = sorted(df["Product"].unique(), key=lambda p: -df.loc[df["Product"] == p, "Sales"].sum())
for i, prod in enumerate(products, start=4):
    ws4.cell(row=i, column=1, value=prod)
    ws4.cell(row=i, column=2, value=f'=SUMIF({rng(PROD_COL)},A{i},{rng(SALES_COL)})')
    ws4.cell(row=i, column=3, value=f'=SUMIF({rng(PROD_COL)},A{i},{rng(PROFIT_COL)})')
    ws4.cell(row=i, column=4, value=f'=C{i}/B{i}')
last4 = 3 + len(products)
style_table_body(ws4, 4, last4, 1, 4)
for r in range(4, last4 + 1):
    ws4.cell(row=r, column=2).number_format = '$#,##0'
    ws4.cell(row=r, column=3).number_format = '$#,##0;($#,##0)'
    ws4.cell(row=r, column=4).number_format = '0.0%'
autosize(ws4, 4, 16)

chart4 = BarChart()
chart4.type = "col"
chart4.title = "Sales vs Profit by Product"
data4 = Reference(ws4, min_col=2, max_col=3, min_row=3, max_row=last4)
cats4 = Reference(ws4, min_col=1, min_row=4, max_row=last4)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.height, chart4.width = 9, 16
ws4.add_chart(chart4, "F3")

# --------------------------------------------------------- 5. Country Summary
ws5 = wb.create_sheet("Country Summary")
ws5["A1"] = "Sales & Margin by Country"
ws5["A1"].font = TITLE_FONT
for j, h in enumerate(["Country", "Sales ($)", "Profit ($)", "Margin (%)"], start=1):
    ws5.cell(row=3, column=j, value=h)
style_table_header(ws5, 3, 1, 4)
countries = sorted(df["Country"].unique(), key=lambda c: -df.loc[df["Country"] == c, "Sales"].sum())
for i, ctry in enumerate(countries, start=4):
    ws5.cell(row=i, column=1, value=ctry)
    ws5.cell(row=i, column=2, value=f'=SUMIF({rng(COUNTRY_COL)},A{i},{rng(SALES_COL)})')
    ws5.cell(row=i, column=3, value=f'=SUMIF({rng(COUNTRY_COL)},A{i},{rng(PROFIT_COL)})')
    ws5.cell(row=i, column=4, value=f'=C{i}/B{i}')
last5 = 3 + len(countries)
style_table_body(ws5, 4, last5, 1, 4)
for r in range(4, last5 + 1):
    ws5.cell(row=r, column=2).number_format = '$#,##0'
    ws5.cell(row=r, column=3).number_format = '$#,##0;($#,##0)'
    ws5.cell(row=r, column=4).number_format = '0.0%'
autosize(ws5, 4, 16)

chart5 = BarChart()
chart5.type = "col"
chart5.title = "Total Sales by Country"
data5 = Reference(ws5, min_col=2, min_row=3, max_row=last5)
cats5 = Reference(ws5, min_col=1, min_row=4, max_row=last5)
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
chart5.height, chart5.width = 9, 16
ws5.add_chart(chart5, "F3")

# --------------------------------------------------------- 6. Monthly Trend
ws6 = wb.create_sheet("Monthly Trend")
ws6["A1"] = "Monthly Sales & Profit Trend"
ws6["A1"].font = TITLE_FONT
for j, h in enumerate(["Year-Month", "Sales ($)", "Profit ($)"], start=1):
    ws6.cell(row=3, column=j, value=h)
style_table_header(ws6, 3, 1, 3)

month_df = (df.assign(YM=df["Year"].astype(str) + "-" + df["Month Number"].astype(int).astype(str).str.zfill(2))
              .sort_values(["Year", "Month Number"])["YM"].unique())
YM_COL_LETTER = None
# Build a helper YM column formula-driven per row isn't needed; use SUMPRODUCT on Year & Month Number
for i, ym in enumerate(month_df, start=4):
    yr, mo = ym.split("-")
    ws6.cell(row=i, column=1, value=ym)
    f_sales = (f'=SUMIFS({rng(SALES_COL)},{rng(YEAR_COL)},{yr},{rng(MONTHNUM_COL)},{int(mo)})')
    f_profit = (f'=SUMIFS({rng(PROFIT_COL)},{rng(YEAR_COL)},{yr},{rng(MONTHNUM_COL)},{int(mo)})')
    ws6.cell(row=i, column=2, value=f_sales)
    ws6.cell(row=i, column=3, value=f_profit)
last6 = 3 + len(month_df)
style_table_body(ws6, 4, last6, 1, 3)
for r in range(4, last6 + 1):
    ws6.cell(row=r, column=2).number_format = '$#,##0'
    ws6.cell(row=r, column=3).number_format = '$#,##0;($#,##0)'
autosize(ws6, 3, 14)

chart6 = LineChart()
chart6.title = "Monthly Sales & Profit Trend"
data6 = Reference(ws6, min_col=2, max_col=3, min_row=3, max_row=last6)
cats6 = Reference(ws6, min_col=1, min_row=4, max_row=last6)
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
for s in chart6.series:
    s.marker.symbol = "circle"
    s.smooth = False
chart6.height, chart6.width = 9, 18
ws6.add_chart(chart6, "F3")

# --------------------------------------------------------------- 7. Dashboard
dash = wb.create_sheet("Dashboard", 0)
dash["A1"] = "Financial Sample — Executive Dashboard"
dash["A1"].font = Font(name=FONT_NAME, bold=True, size=18, color="1F4E78")
dash["A2"] = "Sales, Discounts & Profitability — 700 orders, 5 segments, 6 products, 5 countries, Sep 2013–Dec 2014"
dash["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="555555")

kpi_labels = ["Total Sales", "Total Profit", "Overall Margin", "Loss-Making Segment"]
kpi_formulas = [
    f'=SUM({rng(SALES_COL)})',
    f'=SUM({rng(PROFIT_COL)})',
    f'=SUM({rng(PROFIT_COL)})/SUM({rng(SALES_COL)})',
    '="Enterprise"',
]
for i, (lab, f) in enumerate(zip(kpi_labels, kpi_formulas)):
    col = 1 + i * 2
    cell_lab = dash.cell(row=4, column=col, value=lab)
    cell_lab.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    cell_lab.fill = HEADER_FILL
    cell_lab.alignment = Alignment(horizontal="center")
    dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    cell_val = dash.cell(row=5, column=col, value=f)
    cell_val.font = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
    cell_val.alignment = Alignment(horizontal="center")
    dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
dash["A5"].number_format = '$#,##0,,"M"'
dash["C5"].number_format = '$#,##0,,"M"'
dash["E5"].number_format = '0.0%'
autosize(dash, 8, 14)

# Re-attach copies of the six charts to the dashboard for a single at-a-glance view
from copy import deepcopy
dash.add_chart(deepcopy(chart), "A8")
dash.add_chart(deepcopy(chart2), "A25")
dash.add_chart(deepcopy(chart4), "K8")
dash.add_chart(deepcopy(chart5), "K25")
dash.add_chart(deepcopy(chart6), "A42")

wb.save(OUT_XLSX)
print("Workbook saved:", OUT_XLSX)
